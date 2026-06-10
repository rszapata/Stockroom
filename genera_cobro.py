"""
genera_cobro.py
---------------
Procesa el Excel de ventas de MercadoLibre y genera un reporte de cobro
con el neto a depositar cada 10 días (solo fundas).

Uso:
    python genera_cobro.py <archivo_ml.xlsx> [periodo]

Ejemplos:
    python genera_cobro.py ventas_marzo.xlsx 1     → días 1 al 10
    python genera_cobro.py ventas_marzo.xlsx 2     → días 11 al 20
    python genera_cobro.py ventas_marzo.xlsx 3     → días 21 al 31
    python genera_cobro.py ventas_marzo.xlsx       → todos los días

Fórmula neto por venta:
    neto = ingresos / 1.21 + cargo + costo_fijo
    (cargo y costo_fijo ya vienen negativos en el archivo ML)

Reglas de inclusión/exclusión:
    ✓ Fundas individuales con precio
    ✓ Fundas marcadas "Paquete de varios productos=Sí" que tienen precio propio
    ✓ Paquete de 2 fundas (sin precios individuales) → fórmula sobre el paquete
    ✓ Paquete mixto funda+otro → calcula neto de la funda usando tasas de ventas similares
    ✓ "Venta entregada" / "Cerramos el reclamo y te dimos el dinero" → se incluyen
    ✗ Canceladas, Devoluciones
    ✗ "Reclamo cerrado con reembolso al comprador" → excluida
    ✗ Productos que no son fundas (mallas, soportes, vidrios, etc.)
    ✗ Sub-items sin precio que ya están cubiertos por la fila padre
"""

import sys
import re
import argparse
import json
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from pathlib import Path
import io, sys
# Forzar UTF-8 en stdout para compatibilidad con Windows
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

# -- Configuración ----------------------------------------------
IVA              = 1.21
COSTO_IMPUESTOS  = 307.0   # Costo fiscal fijo por venta (aprox. $307)
PALABRAS_FUNDA   = ['funda']
PALABRAS_EXCLUIR = ['malla', 'soporte', 'vidrio', 'templado', 'film protector',
                    'auricular', 'cargador', 'cable']

# -- Lista blanca de IDs de publicación que SIEMPRE se incluyen --------------
# Agregá acá cualquier producto que no sea funda pero que te corresponda cobrar.
# Formato: 'MLAXXXXXXXXXXX' (el ID de la columna "# de publicación" del Excel ML)
PUBLICACIONES_INCLUIR = {
    'MLA3726855056',   # Cable USB-C
}

PERIODOS = {
    '1': (1,  10, 'Días 1 al 10'),
    '2': (11, 20, 'Días 11 al 20'),
    '3': (21, 31, 'Días 21 al 31'),
}

C = {
    'header_bg': 'FF0D0D10', 'header_fg': 'FFE8FF47',
    'funda_bg':  'FF101810', 'funda_fg':  'FFD0F0D0',
    'paq2_bg':   'FF0A1A0A', 'paq2_fg':   'FFB0FFB0',
    'mixto_bg':  'FF101818', 'mixto_fg':  'FF47FFE8',
    'excl_bg':   'FF1A0808', 'excl_fg':   'FFFF6060',
    'total_bg':  'FF08080F', 'total_fg':  'FFE8FF47',
}


# -- Selector de modo -------------------------------------------
def es_mio(titulo, pub_id, modo):
    """
    modo='fundas': incluye fundas + lista blanca
    modo='otros':  incluye todo lo que NO son fundas ni lista blanca
    """
    es_f  = es_funda(titulo)
    en_wb = pub_id in PUBLICACIONES_INCLUIR
    if modo == 'fundas':
        return es_f or en_wb
    else:  # otros
        return not es_f and not en_wb

# -- Helpers ----------------------------------------------------
def es_funda(titulo):
    t = str(titulo).lower()
    return (any(p in t for p in PALABRAS_FUNDA) and
            not any(p in t for p in PALABRAS_EXCLUIR))

def safe_f(v):
    try:
        f = float(v)
        return f if not pd.isna(f) else 0.0
    except (TypeError, ValueError):
        return 0.0

def tiene_precio(v):
    try:
        return pd.notna(v) and float(v) != 0.0
    except (TypeError, ValueError):
        return False

def calc_neto(ing, cargo, cf):
    try:
        return float(ing) / IVA + float(cargo) + float(cf) - COSTO_IMPUESTOS
    except (TypeError, ValueError):
        return 0.0

def parse_dia(fecha_str):
    try:
        m = re.search(r'(\d+)\s+de', str(fecha_str).lower())
        return int(m.group(1)) if m else None
    except Exception:
        return None

def parse_fecha(fecha_str):
    meses = {'enero':1,'febrero':2,'marzo':3,'abril':4,'mayo':5,'junio':6,
             'julio':7,'agosto':8,'septiembre':9,'octubre':10,'noviembre':11,'diciembre':12}
    try:
        m = re.search(r'(\d+)\s+de\s+(\w+)\s+de\s+(\d{4})', str(fecha_str).lower())
        if m:
            return datetime(int(m.group(3)), meses.get(m.group(2), 1), int(m.group(1)))
    except Exception:
        pass
    return None

def es_valida(estado, desc=''):
    """
    Excluye cancelaciones y devoluciones.
    Excluye reclamos donde el vendedor reembolsó al comprador.
    Incluye "Venta entregada" y reclamos resueltos a favor del vendedor.
    """
    e = str(estado).lower()
    d = str(desc).lower()
    if any(k in e for k in ['cancelad', 'devoluci']): return False
    if 'reembolso al comprador' in e:                 return False
    if 'reembolsaste el dinero' in d:                 return False
    return True

def es_paquete_padre(estado):
    return 'paquete de' in str(estado).lower()

# -- Lectura del archivo ML -------------------------------------
def leer_ml(path):
    """Lee el Excel de ML y devuelve lista de filas como dicts (lectura posicional)."""
    df_raw = pd.read_excel(path, header=None)

    # Encontrar fila de headers
    header_row = None
    for i, row in df_raw.iterrows():
        if str(row[0]).strip() == '# de venta':
            header_row = i
            break
    if header_row is None:
        raise ValueError("No se encontró la fila de encabezados '# de venta'.")

    # Mapear columnas por posición (más robusto que por nombre cuando hay duplicados)
    headers = list(df_raw.iloc[header_row])
    col = {}
    for j, h in enumerate(headers):
        hs = str(h).strip()
        if hs not in col:
            col[hs] = j

    # Índices de columnas clave
    IDX = {
        'id':        col.get('# de venta', 0),
        'fecha':     col.get('Fecha de venta', 1),
        'estado':    col.get('Estado', 2),
        'desc':      col.get('Descripción del estado', 3),
        'paquete':   col.get('Paquete de varios productos', 4),
        'ingresos':  col.get('Ingresos por productos (ARS)', 7),
        'cargo':     col.get('Cargo por venta', 8),
        'costo_fijo':col.get('Costo fijo', 9),
        'titulo':    col.get('Título de la publicación', 24),
        'pub_id':    col.get('# de publicación', 22),
    }

    filas = []
    for i, row in df_raw.iterrows():
        if i <= header_row:
            continue
        filas.append({
            'id':        str(row.iloc[IDX['id']] if IDX['id'] < len(row) else '').strip(),
            'fecha_str': str(row.iloc[IDX['fecha']] if IDX['fecha'] < len(row) else '').strip(),
            'estado':    str(row.iloc[IDX['estado']] if IDX['estado'] < len(row) else '').strip(),
            'desc':      str(row.iloc[IDX['desc']] if IDX['desc'] < len(row) else '').strip(),
            'paquete':   str(row.iloc[IDX['paquete']] if IDX['paquete'] < len(row) else '').strip(),
            'titulo':    str(row.iloc[IDX['titulo']] if IDX['titulo'] < len(row) else '').strip(),
            'ingresos':  row.iloc[IDX['ingresos']] if IDX['ingresos'] < len(row) else None,
            'cargo':     row.iloc[IDX['cargo']] if IDX['cargo'] < len(row) else None,
            'costo_fijo':row.iloc[IDX['costo_fijo']] if IDX['costo_fijo'] < len(row) else None,
            'pub_id':    str(row.iloc[IDX['pub_id']] if IDX['pub_id'] < len(row) else '').strip(),
        })
    return filas

# -- Índice de tasas por producto -------------------------------
def construir_indice_tasas(filas):
    """
    Para cada funda vendida individualmente, registra su precio, cargo y costo_fijo.
    Usado para estimar tasas en paquetes mixtos donde el sub-item no tiene precio.
    Clave: palabras clave del título (primeras 4 palabras significativas).
    """
    indice = {}  # titulo_key → {'ing': ..., 'cargo': ..., 'cf': ...}

    for f in filas:
        if not es_funda(f['titulo']) and f.get('pub_id','') not in PUBLICACIONES_INCLUIR: continue
        if not tiene_precio(f['ingresos']): continue
        if es_paquete_padre(f['estado']): continue
        if not es_valida(f['estado'], f['desc']): continue

        ing   = safe_f(f['ingresos'])
        cargo = safe_f(f['cargo'])
        cf    = safe_f(f['costo_fijo'])
        if ing <= 0: continue

        # Clave: título normalizado (primeras 6 palabras)
        key = ' '.join(f['titulo'].lower().split()[:6])
        if key not in indice:
            indice[key] = []
        indice[key].append({'ing': ing, 'cargo': cargo, 'cf': cf})

    return indice

def buscar_tasas_funda(titulo, indice):
    """
    Busca en el índice las tasas (cargo_rate, costo_fijo) para una funda dada.
    Retorna (cargo_rate, costo_fijo_tipico) o None si no encuentra.
    """
    titulo_words = titulo.lower().split()

    # Buscar por palabras clave del modelo (ej: "samsung s25", "iphone 17", etc.)
    best_match = None
    best_score = 0

    for key, registros in indice.items():
        key_words = key.split()
        score = sum(1 for w in key_words if w in titulo_words and len(w) > 3)
        if score > best_score:
            best_score = score
            best_match = registros

    if best_match and best_score >= 2:
        # Usar el registro más reciente (último en lista)
        ref = best_match[-1]
        cargo_rate = ref['cargo'] / ref['ing'] if ref['ing'] != 0 else -0.16
        cf_tipico  = ref['cf']
        return cargo_rate, cf_tipico

    # Fallback: tasa estándar ML Argentina
    return -0.16, -2925.0

def estimar_neto_funda_en_paquete(titulo_funda, ing_paquete, titulos_sub, indice):
    """
    Para un paquete mixto (funda + otro), estima el neto de la funda.
    Estrategia: calcular precio de la funda como fracción del paquete según tasas históricas.
    """
    cargo_rate, cf_tipico = buscar_tasas_funda(titulo_funda, indice)

    # Estimar precio de la funda: si hay referencias, usar el precio promedio
    key = ' '.join(titulo_funda.lower().split()[:6])
    ref_prices = [r['ing'] for r in indice.get(key, [])]

    if ref_prices:
        # Usar precio de referencia promedio de ventas similares
        ing_funda = sum(ref_prices) / len(ref_prices)
    else:
        # Estimación proporcional: si hay 2 sub-items, funda ≈ mitad
        ing_funda = ing_paquete / len(titulos_sub)

    cargo_funda = ing_funda * cargo_rate
    neto_est = ing_funda / IVA + cargo_funda + cf_tipico - COSTO_IMPUESTOS
    return round(neto_est, 2), round(ing_funda, 2)

# -- Procesamiento principal ------------------------------------
def procesar(filas, modo="fundas", desde_dt=None, hasta_dt=None):
    # Construir índice de tasas usando TODO el archivo (no solo el período)
    indice_tasas = construir_indice_tasas(filas)
    es_modo_fundas = (modo == "fundas")

    ventas = []
    i = 0

    while i < len(filas):
        f = filas[i]
        if not f['id'] or f['id'] == 'nan':
            i += 1
            continue

        dia      = parse_dia(f['fecha_str'])
        fecha_dt = parse_fecha(f['fecha_str'])

        # Filtro por rango de fechas
        if (desde_dt or hasta_dt) and fecha_dt:
            fd = fecha_dt.date() if hasattr(fecha_dt, 'date') else fecha_dt
            if desde_dt and fd < desde_dt:
                i += 1
                continue
            if hasta_dt and fd > hasta_dt:
                i += 1
                continue

        # -- PAQUETE PADRE ---------------------------------------
        if es_paquete_padre(f['estado']):
            sub_sin = []
            sub_con = []
            j = i + 1
            while j < len(filas):
                s = filas[j]
                if s['paquete'] == 'Sí' and s['fecha_str'] == f['fecha_str']:
                    (sub_con if tiene_precio(s['ingresos']) else sub_sin).append(s)
                    j += 1
                else:
                    break

            # Sub-items con precio propio → el padre no aporta datos, se procesan solos
            if sub_con:
                i = j
                continue

            titulos  = [s['titulo'] for s in sub_sin]
            fundas   = [t for t in titulos if es_mio(t, '', modo)]
            otros    = [t for t in titulos if not es_mio(t, '', modo)]
            ing  = safe_f(f['ingresos'])
            cargo_v = safe_f(f['cargo'])
            cf_v = safe_f(f['costo_fijo'])
            tit  = f'PKT: {" + ".join(titulos)}'[:100]

            if not es_valida(f['estado'], f['desc']):
                ventas.append(_v(f, dia, fecha_dt, 'Paquete cancelado',
                                 ing, cargo_v, cf_v, 0.0, excluida=True,
                                 titulo_d=tit, notas='Cancelado/Devolución'))

            elif not titulos or len(fundas) == len(titulos):
                # Todas fundas → fórmula completa
                neto_val = calc_neto(ing, cargo_v, cf_v)
                ventas.append(_v(f, dia, fecha_dt, f'Paquete {len(fundas)} fundas',
                                 ing, cargo_v, cf_v, neto_val,
                                 titulo_d=tit, notas=' + '.join(titulos)))

            elif len(fundas) == 0:
                # Ninguna funda → excluir
                ventas.append(_v(f, dia, fecha_dt, 'Paquete sin fundas',
                                 ing, cargo_v, cf_v, 0.0, excluida=True,
                                 titulo_d=tit, notas=' + '.join(titulos)))

            else:
                # Paquete mixto → calcular neto estimado de la funda
                titulo_funda = fundas[0]
                neto_est, ing_est = estimar_neto_funda_en_paquete(
                    titulo_funda, ing, titulos, indice_tasas)
                notas = (f'FUNDA: {"; ".join(fundas)} | '
                         f'OTROS: {"; ".join(otros)} | '
                         f'Precio funda estimado: ${ing_est:,.0f}')
                ventas.append(_v(f, dia, fecha_dt,
                                 f'Paquete mixto ({len(fundas)}F+{len(otros)}O)',
                                 ing, cargo_v, cf_v, neto_est,
                                 titulo_d=tit, notas=notas, mixto=True))

            i = j
            continue

        # -- SUB-ITEM SIN PRECIO → omitir -----------------------
        if f['paquete'] == 'Sí' and not tiene_precio(f['ingresos']):
            i += 1
            continue

        # -- VENTA INDIVIDUAL ------------------------------------
        ing   = safe_f(f['ingresos'])
        cargo_v = safe_f(f['cargo'])
        cf_v  = safe_f(f['costo_fijo'])

        if not es_valida(f['estado'], f['desc']):
            ventas.append(_v(f, dia, fecha_dt, 'Cancelada/Devolución',
                             ing, cargo_v, cf_v, 0.0,
                             excluida=True, notas='Excluida'))
            i += 1
            continue

        if not es_mio(f['titulo'], f.get('pub_id',''), modo):
            ventas.append(_v(f, dia, fecha_dt, 'No incluido en este modo',
                             ing, cargo_v, cf_v, 0.0,
                             excluida=True, notas='Fuera del modo seleccionado'))
            i += 1
            continue

        if f.get('pub_id','') in PUBLICACIONES_INCLUIR:
            tipo = 'Producto (lista blanca)'
        elif modo == 'otros':
            tipo = 'Producto de socios'
        elif f['paquete'] != 'Sí':
            tipo = 'Funda individual'
        else:
            tipo = 'Funda (en paquete)'
        neto_val = calc_neto(ing, cargo_v, cf_v)
        ventas.append(_v(f, dia, fecha_dt, tipo, ing, cargo_v, cf_v, neto_val))
        i += 1

    return ventas


def _v(f, dia, fecha_dt, tipo, ingresos, cargo, costo_fijo, neto_val,
       excluida=False, mixto=False, notas='', titulo_d=None):
    return {
        'id': f['id'], 'fecha': fecha_dt, 'dia': dia, 'pub_id': f.get('pub_id',''),
        'titulo': titulo_d or f['titulo'],
        'tipo': tipo, 'ingresos': ingresos, 'cargo': cargo,
        'costo_fijo': costo_fijo, 'neto': neto_val,
        'estado': f['estado'], 'notas': notas,
        'excluida': excluida, 'mixto': mixto,
    }

# -- Generación del Excel ---------------------------------------
def generar_excel(ventas, rango_label, output_path, modo="fundas"):
    """rango_label: string descriptivo del período, p.ej. '2025-05-01 al 2025-05-15'."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Cobro Fundas'

    def fl(h):  return PatternFill('solid', start_color=h, end_color=h)
    def fn(color, bold=False, sz=10): return Font(name='Arial', color=color, bold=bold, size=sz)
    def brd():
        s = Side(style='thin', color='FF2a2a30')
        return Border(left=s, right=s, top=s, bottom=s)
    def aln(h='left', w=False): return Alignment(horizontal=h, vertical='center', wrap_text=w)

    p_label = rango_label or 'Período completo'

    # -- Títulos -------------------------------------------------
    ws.merge_cells('A1:I1')
    ws['A1'].value     = f'REPORTE DE COBRO — FUNDAS — {p_label.upper()}'
    ws['A1'].font      = fn(C['header_fg'], bold=True, sz=14)
    ws['A1'].fill      = fl(C['header_bg'])
    ws['A1'].alignment = aln('center')
    ws.row_dimensions[1].height = 32

    ws.merge_cells('A2:I2')
    ws['A2'].value     = (f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}  |  '
                          f'neto = ingresos/1.21 + cargo + costo_fijo − ${COSTO_IMPUESTOS:.0f} (impuestos)  |  '
                          f'Paquetes mixtos: neto calculado con precio estimado de la funda')
    ws['A2'].font      = fn('FF8888a0', sz=9)
    ws['A2'].fill      = fl(C['header_bg'])
    ws['A2'].alignment = aln('center')
    ws.row_dimensions[2].height = 16

    # -- Headers columnas ----------------------------------------
    headers = ['Fecha', '# de Venta', 'Título / Productos', 'Tipo',
               'Ingresos (ARS)', 'Cargo ML', 'Costo Fijo', 'NETO s/IVA (ARS)', 'Notas']
    widths  = [13, 20, 52, 28, 15, 12, 12, 18, 50]
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=3, column=ci, value=h)
        c.font = fn(C['header_fg'], bold=True, sz=10)
        c.fill = fl(C['header_bg'])
        c.border = brd()
        c.alignment = aln('center')
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[3].height = 22
    ws.freeze_panes = 'A4'

    # -- Filas de datos ------------------------------------------
    ROW_START = 4
    fila = ROW_START

    for v in ventas:
        if v['excluida']:
            bg, fg = C['excl_bg'], C['excl_fg']
        elif v['mixto']:
            bg, fg = C['mixto_bg'], C['mixto_fg']
        elif 'Paquete' in v['tipo']:
            bg, fg = C['paq2_bg'], C['paq2_fg']
        else:
            bg, fg = C['funda_bg'], C['funda_fg']

        vals = [
            v['fecha'].strftime('%d/%m/%Y') if v['fecha'] else '—',
            v['id'],
            v['titulo'][:100],
            v['tipo'],
            v['ingresos'],
            v['cargo'],
            v['costo_fijo'],
            v['neto'],
            v['notas'][:150] if v['notas'] else '',
        ]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=fila, column=ci, value=val)
            c.font      = fn(fg, sz=9)
            c.fill      = fl(bg)
            c.border    = brd()
            c.alignment = aln('left', w=(ci in [3, 9]))
            if ci in [5, 6, 7] and isinstance(val, (int, float)):
                c.number_format = '#,##0.00'
            if ci == 8 and isinstance(val, (int, float)):
                c.number_format = '#,##0.00'

        ws.row_dimensions[fila].height = 16
        fila += 1

    # -- Fila TOTAL ----------------------------------------------
    ft = fila + 1
    ws.merge_cells(f'A{ft}:G{ft}')
    c = ws[f'A{ft}']
    c.value = 'TOTAL NETO A DEPOSITAR'
    c.font = fn(C['total_fg'], bold=True, sz=13)
    c.fill = fl(C['total_bg'])
    c.alignment = aln('right')
    c.border = brd()

    c = ws[f'H{ft}']
    c.value         = f'=SUM(H{ROW_START}:H{fila-1})'
    c.font          = fn(C['total_fg'], bold=True, sz=13)
    c.fill          = fl(C['total_bg'])
    c.number_format = '#,##0.00'
    c.border        = brd()
    c.alignment     = aln('right')
    ws[f'I{ft}'].fill   = fl(C['total_bg'])
    ws[f'I{ft}'].border = brd()
    ws.row_dimensions[ft].height = 28

    # -- Leyenda -------------------------------------------------
    fl2 = ft + 2
    ws.cell(row=fl2, column=1, value='LEYENDA').font = fn('FF8888a0', bold=True, sz=9)
    leyendas = [
        ('  Funda individual o paquete de 2 fundas — calculado automáticamente',
         C['funda_bg'], C['funda_fg']),
        ('  Paquete mixto (funda + otro) — neto calculado estimando precio de la funda con ventas similares',
         C['mixto_bg'], C['mixto_fg']),
        ('  Venta excluida (cancelada, devolución, reclamo reembolsado, o no es funda)',
         C['excl_bg'], C['excl_fg']),
    ]
    for off, (txt, bg, fg) in enumerate(leyendas, 1):
        ws.merge_cells(f'A{fl2+off}:I{fl2+off}')
        c = ws.cell(row=fl2+off, column=1, value=txt)
        c.font      = fn(fg, sz=9)
        c.fill      = fl(bg)
        c.alignment = aln('left')

    # -- Hoja Resumen --------------------------------------------
    ws2 = wb.create_sheet('Resumen')
    ws2.column_dimensions['A'].width = 50
    ws2.column_dimensions['B'].width = 22

    incluidas = [v for v in ventas if not v['excluida']]
    excluidas = [v for v in ventas if v['excluida']]
    mixtos    = [v for v in ventas if v['mixto']]
    neto_auto = sum(v['neto'] for v in incluidas)

    rows_res = [
        (f'REPORTE DE COBRO — FUNDAS — {p_label}', '',     True,  13),
        (f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}', '', False, 9),
        ('', '', False, 10),
        ('Ventas de fundas incluidas',                      len(incluidas), False, 10),
        ('  → Fundas individuales',                         len([v for v in incluidas if 'individual' in v['tipo'] or 'paquete' not in v['tipo'].lower()]), False, 10),
        ('  → Paquetes de 2 fundas',                        len([v for v in incluidas if 'Paquete' in v['tipo'] and not v['mixto']]), False, 10),
        ('  → Paquetes mixtos (neto estimado automáticamente)', len(mixtos), False, 10),
        ('Ventas excluidas',                                len(excluidas), False, 10),
        ('', '', False, 10),
        ('TOTAL NETO A DEPOSITAR',       f"='Cobro Fundas'!H{ft}", False, 13),
    ]
    for ri, (lbl, val, bold, sz) in enumerate(rows_res, 1):
        c1 = ws2.cell(row=ri, column=1, value=lbl)
        c1.font = Font(name='Arial', bold=bold, size=sz)
        if val != '':
            c2 = ws2.cell(row=ri, column=2, value=val)
            c2.font = Font(name='Arial', size=sz)
            if isinstance(val, (int, float)):
                c2.number_format = '#,##0.00'
            if isinstance(val, str) and val.startswith('='):
                c2.number_format = '#,##0.00'

    wb.save(output_path)

    # -- Resumen en terminal --------------------------------------
    print(f"\n{'-'*58}")
    print(f"  ✓  {output_path}")
    print(f"{'-'*58}")
    print(f"  Período:                   {p_label}")
    print(f"  Fundas incluidas:          {len(incluidas)}")
    print(f"    → individuales/paquete2: {len(incluidas) - len(mixtos)}")
    print(f"    → paquetes mixtos:       {len(mixtos)}  (neto estimado automáticamente)")
    print(f"  Excluidas:                 {len(excluidas)}")
    print(f"  Impuestos descontados:     ${COSTO_IMPUESTOS:.0f} × {len(incluidas)} ventas = ${COSTO_IMPUESTOS * len(incluidas):>10,.2f}")
    print(f"  TOTAL NETO:                ${neto_auto:>12,.2f}")
    print(f"{'-'*58}\n")
    if mixtos:
        print("  Paquetes mixtos incluidos (verificar estimación):")
        for v in mixtos:
            print(f"    • {v['notas'][:100]}")
            print(f"      Neto estimado: ${v['neto']:,.2f}")
        print()

    # Emitir resumen JSON para el servidor Node
    resumen = {
        'modo':       modo,
        'periodo':    p_label,
        'incluidas':  len(incluidas),
        'mixtos':     len(mixtos),
        'excluidas':  len(excluidas),
        'total_neto': round(neto_auto, 2),
        'mixtos_det': [{'titulo': v['titulo'][:60], 'neto': round(v['neto'],2)} for v in mixtos]
    }
    print(f"RESUMEN_JSON:{json.dumps(resumen)}")

# -- Main -------------------------------------------------------
if __name__ == '__main__':
    ap = argparse.ArgumentParser()
    ap.add_argument('archivo')
    ap.add_argument('--desde',    default=None, help='Fecha inicio YYYY-MM-DD (inclusive)')
    ap.add_argument('--hasta',    default=None, help='Fecha fin   YYYY-MM-DD (inclusive)')
    ap.add_argument('--output',   default=None)
    ap.add_argument('--modo',     default='fundas', choices=['fundas','otros'])
    args = ap.parse_args()

    archivo = args.archivo
    modo    = args.modo

    # Parsear rango de fechas
    desde_dt = hasta_dt = None
    if args.desde:
        try: desde_dt = datetime.strptime(args.desde, '%Y-%m-%d').date()
        except ValueError: pass
    if args.hasta:
        try: hasta_dt = datetime.strptime(args.hasta, '%Y-%m-%d').date()
        except ValueError: pass

    if not Path(archivo).exists():
        print(f"\n  Error: No se encontró '{archivo}'\n")
        sys.exit(1)

    rango_label = f"{args.desde or 'inicio'} al {args.hasta or 'fin'}" if (args.desde or args.hasta) else 'Período completo'
    print(f"\n  Procesando: {archivo}  [modo={modo}]  [{rango_label}]")

    filas  = leer_ml(archivo)
    ventas = procesar(filas, modo=modo, desde_dt=desde_dt, hasta_dt=hasta_dt)

    fecha_hoy = datetime.now().strftime('%Y%m%d')
    p_str     = f"{args.desde}_{args.hasta}" if (args.desde or args.hasta) else 'completo'
    prefix    = 'cobro_fundas' if modo == 'fundas' else 'cobro_otros'
    output    = args.output or f'{prefix}_{p_str}_{fecha_hoy}.xlsx'

    generar_excel(ventas, rango_label, output, modo)
