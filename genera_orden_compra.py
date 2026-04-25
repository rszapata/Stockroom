"""
genera_orden_compra.py — Stockroom
Uso: python genera_orden_compra.py <csv_input> --output <xlsx_output> [--tc 1421] [--flete 800000] [--units 415]
"""
import csv, sys, math, argparse, json
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def main():
    p = argparse.ArgumentParser()
    p.add_argument('csv_input')
    p.add_argument('--output', required=True)
    p.add_argument('--tc',     type=float, default=1421,   help='Tipo de cambio ARS/USD')
    p.add_argument('--flete',  type=float, default=800000, help='Flete + impuestos ARS')
    p.add_argument('--units',  type=int,   default=415,    help='Unidades de referencia para prorrateo')
    p.add_argument('--vendidos-min', type=int, default=5,  help='Umbral mínimo de vendidos para incluir (default 5)')
    p.add_argument('--stock-max',    type=int, default=7,  help='Umbral máximo de stock para incluir (default 7)')
    p.add_argument('--all-products', action='store_true', help='Incluir todas las publicaciones, no solo fundas')
    args = p.parse_args()

    TC         = args.tc
    FLETE_UNIT = round(args.flete / args.units)
    VEND_MIN   = args.vendidos_min
    STOCK_MAX  = args.stock_max

    # Costos USD conocidos por línea de producto (fundas)
    COSTOS_USD = {
        'con_costura':   1.51,
        'samsung_s_cu':  1.51,
        's25_magnetica': 1.60,
        'iphone17_mag':  1.33,
        'magsafe':       0.85,
    }

    # Si contiene estas palabras, NO es funda (productos accesorios o protectores que no son fundas)
    PATRONES_NO_FUNDA = [
        'vidrio templado', 'vidrio protector', 'film protector',
        'protector pantalla', 'protector de pantalla', 'cubre pantalla', 'cubrepantalla',
        'protector lente', 'protector de lente', 'protector camara', 'protector de cámara',
        'protector watch', 'protector para watch', 'protector para samsung watch',
        'protector para apple watch', 'protector reloj', 'protector para reloj',
        'cargador', 'cable', 'auricular', 'auriculares',
        'soporte', 'tripode', 'trípode', 'malla', 'pulsera', 'correa',
        'lente', 'palo selfie', 'powerbank', 'power bank',
        'watch', 'reloj',
    ]
    # Patrones que indican fuertemente que ES una funda
    PATRONES_FUNDA = [
        'funda', 'case', 'carcasa',
        'magsafe', 'magnét', 'magnet',
        'con costura', 'silicona', 'silicone',
    ]
    def es_funda(titulo):
        """True si el producto es una funda (por palabra explícita o patrón típico)."""
        t = titulo.lower()
        if any(p in t for p in PATRONES_NO_FUNDA): return False
        if any(p in t for p in PATRONES_FUNDA): return True
        # Patrón: línea típica del catálogo (S22/S23/S24, S25 Ultra/Premium)
        if 's22' in t and 's23' in t: return True
        if 's25' in t and ('ultra' in t or 'premi' in t): return True
        return False

    def clasificar(titulo):
        """Devuelve clave de costo para la línea de funda."""
        t = titulo.lower()
        if 'con costura' in t and ('iphone' in t or '17' in t): return 'con_costura'
        if 's22 s23 s24' in t:                                   return 'samsung_s_cu'
        if 's25' in t and ('premium' in t or 'premi' in t):      return 'samsung_s_cu'
        if 's25' in t and 'magnét' in t:                         return 's25_magnetica'
        if '17' in t and ('magnét' in t or 'magsafe' in t):      return 'iphone17_mag'
        return 'magsafe'  # iPhone MagSafe u otras fundas

    def costo_unit(titulo):
        """Devuelve costo unitario en ARS."""
        tipo = clasificar(titulo)
        return round(COSTOS_USD[tipo] * TC + FLETE_UNIT)

    def cant_sug(stock, vendidos):
        target = max(10, math.ceil(vendidos * 0.8 / 5) * 5)
        return max(0, target - stock)

    def cant_adj(vendidos):
        return max(5, math.ceil(vendidos / 4 / 5) * 5)

    # ── Leer CSV ────────────────────────────────────────────────
    with open(args.csv_input, 'r', encoding='utf-8-sig') as f:
        rows = list(csv.DictReader(f))

    variantes = [r for r in rows if r['Variante'] != '(total)']
    for r in variantes:
        r['Stock']   = int(float(r['Stock']))
        r['Vendidos']= int(float(r['Vendidos']))
        r['Precio']  = round(float(r['Precio']))

    fundas = variantes if args.all_products else [r for r in variantes if es_funda(r['Título'])]
    # Diagnóstico de filtros
    pass_vend  = sum(1 for r in fundas if r['Vendidos'] >= VEND_MIN)
    pass_stock = sum(1 for r in fundas if r['Stock'] <= STOCK_MAX)
    lista = [r for r in fundas if r['Vendidos'] >= VEND_MIN and r['Stock'] <= STOCK_MAX]

    # Fallback automático: si lista vacía y los umbrales son los default, relajar
    relaxed_used = False
    if not lista and VEND_MIN == 5 and STOCK_MAX == 7:
        VEND_MIN_R, STOCK_MAX_R = 1, 15
        lista = [r for r in fundas if r['Vendidos'] >= VEND_MIN_R and r['Stock'] <= STOCK_MAX_R]
        if lista:
            VEND_MIN, STOCK_MAX = VEND_MIN_R, STOCK_MAX_R
            relaxed_used = True

    # Ordenar por publicación luego por vendidos desc
    orden = [
        "iPhone 17 Pro Max Air | Con Costura",
        "iPhone 17 Pro Max Air Magnética",
        "iPhone 17 Pro Max Air Magsafe",
        "iPhone 16/15/14/13/12 | Con Costura",
        "iPhone 16/15/14/13/12 Magnética Magsafe",
        "S25 Ultra | Magnética",
        "S25 Ultra | Premium",
        "S22 S23 S24 Ultra",
    ]
    def sort_key(r):
        t = r['Título']
        for i, o in enumerate(orden):
            if o.lower() in t.lower(): return (i, -r['Vendidos'], r['Stock'])
        return (99, -r['Vendidos'], r['Stock'])

    lista.sort(key=sort_key)

    # ── Estilos ──────────────────────────────────────────────────
    thin = Side(style='thin', color='BDD7EE')
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)
    C_DB="1F3864"; C_BL="2E75B6"; C_WH="FFFFFF"
    C_CR="FFD7D7"; C_LW="FFF2CC"; C_AL="EEF3FB"
    C_GB="E2EFDA"; C_GH="375623"
    C_OR_BG="FFF0E0"; C_OR_FT="BF5700"

    def hdr(c, bg=C_DB, fc=C_WH, sz=9):
        c.font=Font(name='Arial',bold=True,color=fc,size=sz)
        c.fill=PatternFill('solid',start_color=bg)
        c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
        c.border=brd

    def dat(c, bg=C_WH, bold=False, align='center', sz=9, col="000000"):
        c.font=Font(name='Arial',size=sz,bold=bold,color=col)
        c.fill=PatternFill('solid',start_color=bg)
        c.alignment=Alignment(horizontal=align,vertical='center',wrap_text=True)
        c.border=brd

    wb = Workbook()

    # ══ HOJA 1: ORDEN DE COMPRA ═════════════════════════════════
    ws = wb.active
    ws.title = "🧾 Orden de Compra"
    ws.sheet_view.showGridLines = False

    ws.merge_cells('A1:L1')
    ws['A1'] = "🧾 ORDEN DE COMPRA — Fundas con Alta Rotación y Stock Bajo"
    ws['A1'].font = Font(name='Arial',bold=True,size=14,color=C_WH)
    ws['A1'].fill = PatternFill('solid',start_color=C_DB)
    ws['A1'].alignment = Alignment(horizontal='center',vertical='center')
    ws.row_dimensions[1].height = 32

    ws.merge_cells('A2:L2')
    extra = "  |  ⚠ umbrales relajados automáticamente (no había items con criterio estándar)" if relaxed_used else ""
    ws['A2'] = f"Criterio: Vendidos ≥ {VEND_MIN}  |  Stock ≤ {STOCK_MAX}  |  TC: ${TC:,.0f}  |  Flete/u: ${FLETE_UNIT:,}{extra}"
    ws['A2'].font = Font(name='Arial',italic=True,size=8,color="444444")
    ws['A2'].fill = PatternFill('solid',start_color="DCE6F1")
    ws['A2'].alignment = Alignment(horizontal='center',vertical='center')
    ws.row_dimensions[2].height = 16

    hdrs = ["#","Prioridad","Producto","Color","Modelo","Stock\nActual","Vendidas","Cant.\nSugerida","Cant.\nAjustada","Costo\nUnit. $","Costo\nTotal $","Total\nAjustado $"]
    ws.row_dimensions[3].height = 34
    for i,h in enumerate(hdrs,1): hdr(ws.cell(row=3,column=i,value=h))

    widths=[4,12,36,12,22,8,9,10,10,12,14,14]
    for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w

    titulo_ant = None
    row_actual = 4
    num = 1
    total_sug=0; total_adj=0; total_costo_sug=0; total_costo_adj=0

    for r in lista:
        titulo = r['Título']
        if titulo != titulo_ant:
            ws.row_dimensions[row_actual].height = 18
            ws.merge_cells(f'A{row_actual}:L{row_actual}')
            c = ws.cell(row=row_actual,column=1,value=f"📦  {titulo}")
            c.font=Font(name='Arial',bold=True,size=9,color=C_WH)
            c.fill=PatternFill('solid',start_color=C_BL)
            c.alignment=Alignment(horizontal='left',vertical='center')
            c.border=brd
            row_actual+=1; titulo_ant=titulo

        ws.row_dimensions[row_actual].height = 22
        stock = r['Stock']
        if stock == 0:   bp="FFB3B3"
        elif stock <= 4: bp=C_CR
        else:            bp=C_LW
        ab = C_AL if num%2==0 else C_WH

        variante = r['Variante']
        color=""; modelo=variante
        if '·' in variante:
            pts = variante.split('·')
            color = pts[0].replace('Color:','').strip()
            modelo = pts[1].replace('Nombre del diseño:','').strip() if len(pts)>1 else ''

        cu = costo_unit(titulo)
        csug = cant_sug(stock, r['Vendidos'])
        cadj = cant_adj(r['Vendidos'])

        # Color de costo según línea
        if 'costura' in titulo.lower() or 's22' in titulo.lower() or 's25' in titulo.lower():
            bg_cost, ft_cost = "E2EFDA", "375623"
        elif '17' in titulo.lower():
            bg_cost, ft_cost = "E3F2FD", "0D47A1"
        else:
            bg_cost, ft_cost = "E8F5E9", "1B5E20"

        total_sug_val = csug * cu
        total_adj_val = cadj * cu
        bg_adj_cost   = C_OR_BG
        ft_adj_cost   = C_OR_FT

        for ci,(val,align,bg,bold,colf) in enumerate([
            (num,'center',ab,False,"000000"),
            ("🔴 CRÍTICO" if stock<=4 else "🟡 BAJO",'center',bp,True,"000000"),
            (titulo,'left',ab,False,"000000"),
            (color,'center',ab,False,"000000"),
            (modelo,'left',ab,False,"000000"),
            (stock,'center',bp,True,"7B0000" if stock==0 else "000000"),
            (r['Vendidos'],'center',ab,True,C_DB),
            (csug,'center',C_GB,True,C_GH),
            (cadj,'center',C_OR_BG,True,C_OR_FT),
            (cu,'center',bg_cost,False,ft_cost),
            (total_sug_val,'center',bg_cost,True,ft_cost),
            (total_adj_val,'center',bg_adj_cost,True,ft_adj_cost),
        ],1):
            c=ws.cell(row=row_actual,column=ci,value=val)
            dat(c,bg=bg,bold=bold,align=align,col=colf)
            if ci in (10,11,12) and isinstance(val, int): c.number_format='$#,##0'

        total_sug+=csug; total_adj+=cadj
        total_costo_sug+=csug*cu; total_costo_adj+=cadj*cu
        row_actual+=1; num+=1

    # Fila TOTAL
    tr = row_actual
    ws.row_dimensions[tr].height = 26
    ws.merge_cells(f'A{tr}:G{tr}')
    c=ws.cell(row=tr,column=1,value="TOTAL ESTIMADO DE INVERSIÓN")
    c.font=Font(name='Arial',bold=True,size=10,color=C_WH)
    c.fill=PatternFill('solid',start_color=C_DB)
    c.alignment=Alignment(horizontal='right',vertical='center')
    c.border=brd
    for ci,val,fmt in [(8,total_sug,None),(9,total_adj,None),(11,total_costo_sug,'$#,##0'),(12,total_costo_adj,'$#,##0')]:
        c=ws.cell(row=tr,column=ci,value=val)
        dat(c,bg="C6EFCE",bold=True,col=C_GH,sz=11)
        if fmt: c.number_format=fmt
    ws.cell(row=tr,column=10).fill=PatternFill('solid',start_color="C6EFCE")
    ws.cell(row=tr,column=10).border=brd

    nota_row = tr+1
    ws.merge_cells(f'A{nota_row}:L{nota_row}')
    cusd_str=" | ".join([f"USD{v}→${round(v*TC+FLETE_UNIT):,}" for v in sorted(set(COSTOS_USD.values()))])
    c=ws.cell(row=nota_row,column=1,value=f"📌 Costos: {cusd_str}  |  Flete: ${FLETE_UNIT:,}/u  |  TC: ${TC:,.0f}")
    c.font=Font(name='Arial',italic=True,size=8,color="444444")
    c.fill=PatternFill('solid',start_color="F2F2F2")
    c.alignment=Alignment(horizontal='center',vertical='center')
    ws.row_dimensions[nota_row].height=16

    # ══ HOJA 2: RESUMEN ═════════════════════════════════════════
    ws2 = wb.create_sheet("📊 Resumen")
    ws2.sheet_view.showGridLines = False
    ws2.merge_cells('A1:D1')
    ws2['A1']="RESUMEN EJECUTIVO — Inventario MercadoLibre"
    ws2['A1'].font=Font(name='Arial',bold=True,size=13,color=C_WH)
    ws2['A1'].fill=PatternFill('solid',start_color="2E75B6")
    ws2['A1'].alignment=Alignment(horizontal='center',vertical='center')
    ws2.row_dimensions[1].height=28

    kpis=[
        ("Total fundas en catálogo", len(fundas)),
        ("Fundas con stock CRÍTICO (≤4)", sum(1 for r in fundas if r['Stock']<=4)),
        ("Fundas SIN stock (0 unidades)", sum(1 for r in fundas if r['Stock']==0)),
        ("Fundas a reponer (alta rot. + stock bajo)", len(lista)),
    ]
    C_GR="E2EFDA"
    for i,(lbl,val) in enumerate(kpis,1):
        row=i+2; ws2.row_dimensions[row].height=22
        bg="FFD7D7" if "CRÍTICO" in lbl or "SIN stock" in lbl else C_GR
        for col in range(1,5):
            c=ws2.cell(row=row,column=col); c.fill=PatternFill('solid',start_color=bg); c.border=brd
        ws2.cell(row=row,column=1,value=i).font=Font(name='Arial',bold=True,size=9)
        ws2.cell(row=row,column=1).alignment=Alignment(horizontal='center')
        ws2.cell(row=row,column=2,value=lbl).font=Font(name='Arial',size=9)
        v=ws2.cell(row=row,column=3,value=val)
        v.font=Font(name='Arial',bold=True,size=11,color=C_DB); v.alignment=Alignment(horizontal='center')

    ws2.column_dimensions['A'].width=5; ws2.column_dimensions['B'].width=42
    ws2.column_dimensions['C'].width=18; ws2.column_dimensions['D'].width=10

    by_prod=defaultdict(lambda:{'v':0,'c':0})
    for r in lista:
        by_prod[r['Título']]['v']+=1; by_prod[r['Título']]['c']+=r['Vendidos']
    ws2.row_dimensions[8].height=28
    ws2.merge_cells('A8:D8')
    c=ws2.cell(row=8,column=1,value="TOP PRODUCTOS A REPONER (por ventas)")
    c.font=Font(name='Arial',bold=True,size=10,color=C_WH)
    c.fill=PatternFill('solid',start_color=C_DB); c.alignment=Alignment(horizontal='center',vertical='center')
    for i,h in enumerate(["Producto","Variantes","Ventas Totales",""],1): hdr(ws2.cell(row=9,column=i,value=h),bg="2E75B6")
    ws2.row_dimensions[9].height=28
    for i,(prod,d) in enumerate(sorted(by_prod.items(),key=lambda x:-x[1]['c']),1):
        row=9+i; ws2.row_dimensions[row].height=20
        bg=C_AL if i%2==0 else C_WH
        dat(ws2.cell(row=row,column=1,value=prod),bg=bg,align='left')
        dat(ws2.cell(row=row,column=2,value=d['v']),bg=bg,bold=True)
        c=ws2.cell(row=row,column=3,value=d['c']); dat(c,bg=bg,bold=True,col=C_DB)
        ws2.cell(row=row,column=4).fill=PatternFill('solid',start_color=bg); ws2.cell(row=row,column=4).border=brd

    wb.save(args.output)

    # Si la lista quedó vacía, escribir un mensaje informativo en la hoja
    if not lista:
        info_row = 4
        ws.merge_cells(f'A{info_row}:L{info_row}')
        c = ws.cell(row=info_row, column=1,
            value=f"⚠ Ningún item cumple los criterios (Vendidos≥{VEND_MIN}, Stock≤{STOCK_MAX}). "
                  f"De {len(fundas)} {'productos' if args.all_products else 'fundas'} analizados, "
                  f"{pass_vend} cumplen el umbral de vendidos y {pass_stock} el de stock. "
                  f"Probá bajar los umbrales o activar 'todos los productos'.")
        c.font = Font(name='Arial', italic=True, size=10, color='B85450')
        c.fill = PatternFill('solid', start_color='FFF8E1')
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = brd
        ws.row_dimensions[info_row].height = 60

    resumen = {
        "items": len(lista),
        "total_sug": total_sug,
        "total_adj": total_adj,
        "inversion_sug": total_costo_sug,
        "inversion_adj": total_costo_adj,
        "tc": TC,
        "flete_unit": FLETE_UNIT,
        # Diagnóstico para el frontend
        "variantes_total": len(variantes),
        "fundas_total": len(fundas),
        "pass_vendidos": pass_vend,
        "pass_stock": pass_stock,
        "vendidos_min": VEND_MIN,
        "stock_max": STOCK_MAX,
        "all_products": bool(args.all_products),
        "relaxed_used": relaxed_used,
    }
    print(f"RESUMEN_JSON:{json.dumps(resumen)}")

if __name__ == '__main__':
    main()
