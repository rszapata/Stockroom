"""
flex_cost.py
------------
Lee el Excel de ventas de MercadoLibre y calcula el costo total de los envíos
Flex del período (filtra por columna "Forma de entrega" = "Mercado Envíos Flex").
"""
import sys, os, json, argparse, re, unicodedata
from pathlib import Path
from datetime import datetime

import openpyxl

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

TARIFFS = {'caba': 4490, 'gba_cerca': 6490, 'gba_lejos': 8490, 'sin_zona': 0}

ESTADOS_EXCLUIR = {'cancelada', 'cancelado', 'reclamo cerrado con reembolso al comprador'}

def auto_zone(cp):
    if not cp: return None
    s = str(cp).strip().upper()
    if re.match(r'^C\d{4}', s): return 'caba'
    if s.isdigit():
        n = int(s)
        if 1000 <= n <= 1499: return 'caba'
    return None

MESES_ES = {
    'enero':1,'febrero':2,'marzo':3,'abril':4,'mayo':5,'junio':6,
    'julio':7,'agosto':8,'septiembre':9,'setiembre':9,'octubre':10,
    'noviembre':11,'diciembre':12
}

def parse_date(v):
    if v is None: return None
    # openpyxl devuelve datetime directamente para celdas de fecha bien formateadas
    if isinstance(v, datetime): return v
    # openpyxl puede devolver date (sin hora) para celdas sin componente de tiempo
    try:
        from datetime import date as _date
        if type(v) is _date:
            return datetime(v.year, v.month, v.day)
    except Exception: pass
    # Número serial de Excel (openpyxl a veces devuelve float si la celda no tiene
    # formato de fecha reconocido, p.ej. columnas calculadas)
    if isinstance(v, (int, float)):
        try:
            import datetime as _dt
            # Excel epoch: 1 = 1900-01-01; corrección por bug del año bisiesto 1900
            n = int(v)
            if n > 59: n -= 1          # saltar el falso 29/02/1900 de Excel
            delta = _dt.timedelta(days=n - 1)
            base  = _dt.datetime(1900, 1, 1)
            return base + delta
        except Exception: pass
    s = str(v).strip().lower().replace('hs.', '').replace(' hs', '').strip()
    m = re.match(r'^(\d{1,2})\s+de\s+(\w+)\s+de\s+(\d{4})(?:\s+(\d{1,2}):(\d{2}))?', s)
    if m:
        d = int(m.group(1)); mes_name = m.group(2); y = int(m.group(3))
        mes = MESES_ES.get(mes_name)
        if mes:
            hh = int(m.group(4) or 0); mm = int(m.group(5) or 0)
            try: return datetime(y, mes, d, hh, mm)
            except: pass
    for fmt in ('%d/%m/%Y %H:%M hs', '%d/%m/%Y %H:%M', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S', '%Y-%m-%d'):
        try: return datetime.strptime(str(v).strip(), fmt)
        except: pass
    return None

def find_header_row(ws):
    for r in range(1, 15):
        v = ws.cell(r, 1).value
        if v and '# de venta' in str(v).lower():
            return r
    return 6

def normalize_str(s):
    """Quita acentos y pasa a minúsculas para evitar errores si ML cambia una letra."""
    s = str(s).strip().lower()
    return unicodedata.normalize('NFKD', s).encode('ASCII', 'ignore').decode('utf-8')

def get_column_mapping(ws, header_row):
    """
    Construye un mapa {campo: [col1, col2, ...]} escaneando la fila de headers.
    Normaliza los nombres (sin acentos, minúsculas) para tolerar diferencias entre
    el Excel de WZ (tienda oficial) y RZ (vendedor regular), que difieren en posición
    y a veces en el nombre exacto de las columnas.
    """
    col_map = {}
    header_debug = {}  # col_index → nombre raw (para diagnóstico)
    for c in range(1, ws.max_column + 1):
        raw_name = ws.cell(header_row, c).value
        if not raw_name: continue
        header_debug[c] = str(raw_name).strip()
        name = normalize_str(raw_name)
        if name not in col_map:
            col_map[name] = []
        col_map[name].append(c)

    # ── Columnas individuales (primer match o default) ──────────
    mapping = {
        'VENTA':  col_map.get('# de venta',           [1])[0],
        'FECHA':  col_map.get('fecha de venta',        [2])[0],
        'DESC':   col_map.get('descripcion del estado',[4])[0],
        'CP':     col_map.get('codigo postal',         [40])[0],
        'CIUDAD': col_map.get('ciudad',                [38])[0],
        'DOM':    col_map.get('domicilio',             [37])[0],
        # ── Columnas múltiples (se busca Flex en todas) ────────
        # WZ tienda oficial tiene "Forma de entrega" en col ~49; RZ regular en ~42.
        # Al buscar por nombre no importa la posición; el fallback cubre ambas.
        'FORMAS':  col_map.get('forma de entrega',     [42, 49]),
        # Hay dos columnas "Estado": col 3 = status de venta, col ~39 = estado/provincia
        'ESTADOS': col_map.get('estado',               [3, 39]),
        # ── Provincia explícita (solo tienda oficial la tiene) ──
        # Si existe, se usa directamente; si no, se cae al segundo "Estado"
        'PROV': (col_map.get('provincia') or
                 col_map.get('estado (provincia)') or
                 None),
        # ── Meta ──────────────────────────────────────────────
        '_headers': header_debug,
        '_col_map': {k: v for k, v in col_map.items()},
    }
    return mapping

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('excel')
    ap.add_argument('--desde',  default='', help='Fecha inicio YYYY-MM-DD (inclusive)')
    ap.add_argument('--hasta',  default='', help='Fecha fin   YYYY-MM-DD (inclusive)')
    ap.add_argument('--zones',  default='')
    args = ap.parse_args()

    zones_map = {}
    if args.zones and os.path.exists(args.zones):
        try:
            with open(args.zones, 'r', encoding='utf-8') as f:
                zones_map = json.load(f)
        except Exception as e:
            pass

    # Parsear rango de fechas
    from datetime import date as date_cls
    desde_dt = hasta_dt = None
    if args.desde:
        try: desde_dt = datetime.strptime(args.desde, '%Y-%m-%d').date()
        except: pass
    if args.hasta:
        try: hasta_dt = datetime.strptime(args.hasta, '%Y-%m-%d').date()
        except: pass

    wb = openpyxl.load_workbook(args.excel, data_only=True)
    ws = wb.active
    header_row = find_header_row(ws)
    cols = get_column_mapping(ws, header_row)
    data_start = header_row + 1

    breakdown = {}
    unmapped  = {}
    flex_total = 0
    skipped_period = 0
    skipped_estado = 0
    rows_seen = 0
    sample_dates = []

    for r in range(data_start, ws.max_row + 1):
        venta = ws.cell(r, cols['VENTA']).value
        if not venta: continue
        rows_seen += 1
        
        # Busca la palabra 'flex' en cualquiera de las columnas 'Forma de entrega'
        is_flex = False
        for f_col in cols['FORMAS']:
            if 'flex' in str(ws.cell(r, f_col).value or '').lower():
                is_flex = True
                break
        
        if not is_flex: continue

        # Status de la venta (primera columna "Estado")
        estado = ws.cell(r, cols['ESTADOS'][0]).value or ''
        desc   = ws.cell(r, cols['DESC']).value or ''
        st_lower = str(estado).strip().lower()
        ds_lower = str(desc).strip().lower()

        # Filtro de cancelados reales
        if st_lower in ESTADOS_EXCLUIR or ds_lower in ESTADOS_EXCLUIR:
            skipped_estado += 1
            continue
        
        # Evita el falso positivo de "cancelamos la devolución" asegurándose que solo descarte ventas no entregadas
        combined = (st_lower + ' ' + ds_lower)
        if 'venta cancelada' in combined or ('devuelto' in combined and 'entregada' not in combined):
            skipped_estado += 1
            continue

        # Parsear fecha siempre (para mostrar el rango real del archivo en el output)
        fecha = parse_date(ws.cell(r, cols['FECHA']).value)
        if fecha:
            sample_dates.append(fecha.strftime('%Y-%m-%d'))

        # Filtro por rango de fechas
        if desde_dt or hasta_dt:
            if not fecha:
                # Sin fecha parseable → excluir cuando hay filtro activo
                skipped_period += 1
                continue
            fd = fecha.date()
            if desde_dt and fd < desde_dt:
                skipped_period += 1
                continue
            if hasta_dt and fd > hasta_dt:
                skipped_period += 1
                continue

        cp = str(ws.cell(r, cols['CP']).value or '').strip()
        ciudad = ws.cell(r, cols['CIUDAD']).value or ''
        dom    = ws.cell(r, cols['DOM']).value    or ''

        # Provincia: columna explícita (tienda oficial) > segundo "Estado" (regular) > vacío
        if cols['PROV']:
            prov_val = ws.cell(r, cols['PROV'][0]).value or ''
        elif len(cols['ESTADOS']) > 1:
            prov_val = ws.cell(r, cols['ESTADOS'][-1]).value or ''
        else:
            prov_val = ''

        addr = ' · '.join(str(x) for x in [dom, ciudad, prov_val] if x)

        flex_total += 1
        zone = zones_map.get(cp) or auto_zone(cp)
        if zone and zone in TARIFFS:
            entry = breakdown.setdefault(cp, {
                'count': 0, 'sample_address': addr, 'zone': zone,
                'sample_venta': str(venta),
            })
            entry['count'] += 1
            if not entry['sample_address']: entry['sample_address'] = addr
        else:
            entry = unmapped.setdefault(cp, {
                'count': 0, 'sample_address': addr, 'sample_venta': str(venta),
            })
            entry['count'] += 1
            if not entry['sample_address']: entry['sample_address'] = addr

    breakdown_arr = []
    for cp, v in breakdown.items():
        tariff = TARIFFS[v['zone']]
        breakdown_arr.append({
            'cp': cp, 'count': v['count'], 'zone': v['zone'],
            'tariff': tariff, 'subtotal': v['count'] * tariff,
            'sample_address': v['sample_address'],
            'sample_venta': v['sample_venta'],
        })
    breakdown_arr.sort(key=lambda x: -x['subtotal'])

    unmapped_arr = []
    for cp, v in unmapped.items():
        unmapped_arr.append({
            'cp': cp, 'count': v['count'],
            'sample_address': v['sample_address'],
            'sample_venta': v['sample_venta'],
            'auto_suggest': auto_zone(cp),
        })
    unmapped_arr.sort(key=lambda x: -x['count'])

    total_cost = sum(b['subtotal'] for b in breakdown_arr)

    # Resumen de columnas detectadas (útil para debuggear Excels de distintas cuentas)
    detected_cols = {
        'venta':    cols['VENTA'],
        'fecha':    cols['FECHA'],
        'desc':     cols['DESC'],
        'cp':       cols['CP'],
        'ciudad':   cols['CIUDAD'],
        'dom':      cols['DOM'],
        'formas':   cols['FORMAS'],
        'estados':  cols['ESTADOS'],
        'prov':     cols['PROV'],
        'total_headers': len(cols.get('_headers', {})),
    }

    out = {
        'ok': True,
        'flex_shipments': flex_total,
        'mapped_count': sum(b['count'] for b in breakdown_arr),
        'unmapped_count': sum(u['count'] for u in unmapped_arr),
        'total_cost': total_cost,
        'tariffs': TARIFFS,
        'breakdown': breakdown_arr,
        'unmapped': unmapped_arr,
        'periodo': f"{args.desde} → {args.hasta}" if (args.desde or args.hasta) else 'todo',
        'rows_seen': rows_seen,
        'skipped_period': skipped_period,
        'skipped_estado': skipped_estado,
        'date_range': [min(sample_dates), max(sample_dates)] if sample_dates else None,
        'detected_cols': detected_cols,
        '_filtro': {
            'desde_arg': args.desde or None,
            'hasta_arg':  args.hasta  or None,
            'desde_dt':  str(desde_dt)  if desde_dt  else None,
            'hasta_dt':  str(hasta_dt)  if hasta_dt  else None,
        },
    }
    print('FLEX_JSON:' + json.dumps(out, ensure_ascii=False))

if __name__ == '__main__':
    try:
        main()
    except Exception as e:
        print('FLEX_ERROR:' + json.dumps({'error': str(e)}))
        sys.exit(1)